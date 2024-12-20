from openpyxl import load_workbook, Workbook

def load_inventory(excel_file):
    try:
        workbook = load_workbook(excel_file)
        sheet = workbook.active
        inventory = {}
        for row in sheet.iter_rows(min_row=2, values_only=True):
            product_name, quantity = row
            inventory[product_name] = quantity
        return inventory
    except FileNotFoundError:
        return {}

def save_inventory(excel_file, inventory):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Produto", "Quantidade"])
    for product_name, quantity in inventory.items():
        sheet.append([product_name, quantity])
    workbook.save(excel_file)